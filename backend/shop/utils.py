from __future__ import annotations

import io
import json
import mimetypes
import re
import shutil
import urllib.parse
import urllib.request
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from uuid import uuid4

from django.conf import settings
from django.core.files.uploadedfile import UploadedFile
from django.db.models import Model
from openpyxl import load_workbook


FLAT_IMPORT_HEADERS = {
    "title",
    "sku",
    "category_slug",
    "category_image",
    "category_image_url",
    "price",
    "thickness",
    "color",
    "material",
    "description",
    "is_published",
}

PRODUCT_IMAGE_HEADER_RE = re.compile(
    r"^(?:image|image_url|photo|photo_url|img|product_image|product_image_url|product_photo)(\d+)?$"
)
CATEGORY_IMAGE_HEADERS = {
    "category_image",
    "category_image_url",
    "category_photo",
    "category_photo_url",
}


@dataclass
class CatalogImportNode:
    title: str
    depth: int
    sort: int
    media_urls: list[str] = field(default_factory=list)
    children: list["CatalogImportNode"] = field(default_factory=list)


@dataclass
class CatalogImportSheet:
    title: str
    nodes: list[CatalogImportNode]


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-zA-Z0-9\u0400-\u04FF]+", "-", value.strip().lower())
    return slug.strip("-") or uuid4().hex[:12]


def make_unique_slug(model: type[Model], base: str, *, instance_pk=None) -> str:
    base_slug = slugify(base)
    slug = base_slug
    suffix = 1

    while True:
        queryset = model.objects.filter(slug=slug)
        if instance_pk is not None:
            queryset = queryset.exclude(pk=instance_pk)
        if not queryset.exists():
            return slug
        suffix += 1
        slug = f"{base_slug}-{suffix}"


def build_media_url(request, image_path: str | None) -> str | None:
    if not image_path:
        return None
    if image_path.startswith("http://") or image_path.startswith("https://"):
        return image_path
    return request.build_absolute_uri(f"{settings.MEDIA_URL}{image_path.lstrip('/')}")


def ensure_media_root() -> None:
    Path(settings.MEDIA_ROOT).mkdir(parents=True, exist_ok=True)


def remove_media_directory(*parts: str) -> None:
    target_dir = Path(settings.MEDIA_ROOT).joinpath(*parts)
    if target_dir.exists():
        shutil.rmtree(target_dir)


def remove_media_file(path_str: str | None) -> None:
    if not path_str:
        return

    target_path = Path(settings.MEDIA_ROOT) / path_str
    if target_path.exists():
        target_path.unlink()


def store_uploaded_file(scope: str, entity_id: str, file: UploadedFile, *, prefix: str) -> str:
    ensure_media_root()
    entity_dir = Path(settings.MEDIA_ROOT) / scope / entity_id
    entity_dir.mkdir(parents=True, exist_ok=True)

    extension = Path(file.name).suffix.lower() or ".jpg"
    filename = f"{prefix}-{uuid4().hex}{extension}"
    target_path = entity_dir / filename
    with target_path.open("wb") as output:
        for chunk in file.chunks():
            output.write(chunk)
    return str(Path(scope) / entity_id / filename).replace("\\", "/")


def remove_product_media(product_id: str) -> None:
    remove_media_directory("products", product_id)


def remove_category_media(category_id: str) -> None:
    remove_media_directory("categories", category_id)


def replace_product_images(product_id: str, files: list[UploadedFile]) -> list[str]:
    ensure_media_root()
    remove_product_media(product_id)

    product_dir = Path(settings.MEDIA_ROOT) / "products" / product_id
    product_dir.mkdir(parents=True, exist_ok=True)

    stored_paths: list[str] = []
    for index, file in enumerate(files):
        extension = Path(file.name).suffix.lower() or ".jpg"
        filename = f"image-{index + 1}-{uuid4().hex}{extension}"
        target_path = product_dir / filename
        with target_path.open("wb") as output:
            for chunk in file.chunks():
                output.write(chunk)
        stored_paths.append(str(Path("products") / product_id / filename).replace("\\", "/"))

    return stored_paths


def append_product_images(product_id: str, files: list[UploadedFile], *, start_sort: int = 0) -> list[str]:
    ensure_media_root()

    product_dir = Path(settings.MEDIA_ROOT) / "products" / product_id
    product_dir.mkdir(parents=True, exist_ok=True)

    stored_paths: list[str] = []
    for offset, file in enumerate(files):
        extension = Path(file.name).suffix.lower() or ".jpg"
        filename = f"image-{start_sort + offset + 1}-{uuid4().hex}{extension}"
        target_path = product_dir / filename
        with target_path.open("wb") as output:
            for chunk in file.chunks():
                output.write(chunk)
        stored_paths.append(str(Path("products") / product_id / filename).replace("\\", "/"))

    return stored_paths


def guess_media_extension(source_url: str, headers=None) -> str:
    if headers is not None:
        content_type = ""
        if hasattr(headers, "get_content_type"):
            content_type = headers.get_content_type() or ""
        elif hasattr(headers, "get"):
            content_type = headers.get("Content-Type") or ""
        if content_type:
            guessed = mimetypes.guess_extension(content_type.split(";")[0].strip())
            if guessed and guessed.lower() != ".jpe":
                return guessed.lower()
            if guessed and guessed.lower() == ".jpe":
                return ".jpg"

    parsed = urllib.parse.urlparse(source_url)
    suffix = Path(parsed.path).suffix.lower()
    if suffix in {".jpg", ".jpeg", ".png", ".webp", ".gif", ".avif", ".svg"}:
        return ".jpg" if suffix == ".jpeg" else suffix
    return ".jpg"


def is_remote_media_url(value: str | None) -> bool:
    text = normalize_media_reference(value)
    parsed = urllib.parse.urlparse(text)
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def download_remote_media(scope: str, entity_id: str, source_url: str, *, prefix: str) -> str:
    ensure_media_root()
    entity_dir = Path(settings.MEDIA_ROOT) / scope / entity_id
    entity_dir.mkdir(parents=True, exist_ok=True)

    request = urllib.request.Request(
        normalize_media_reference(source_url),
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0 Safari/537.36"
            ),
            "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
            "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
        },
    )
    with urllib.request.urlopen(request, timeout=30) as response:
        extension = guess_media_extension(source_url, response.headers)
        filename = f"{prefix}-{uuid4().hex}{extension}"
        target_path = entity_dir / filename
        with target_path.open("wb") as output:
            shutil.copyfileobj(response, output)
    return str(Path(scope) / entity_id / filename).replace("\\", "/")


def materialize_category_image(category_id: str, image_source: str | None) -> str | None:
    normalized = normalize_media_reference(image_source)
    if not normalized:
        return None
    if is_remote_media_url(normalized):
        remove_category_media(category_id)
        return download_remote_media("categories", category_id, normalized, prefix="cover")
    return normalized


def materialize_product_images(product_id: str, image_sources: list[str]) -> list[str]:
    normalized_sources = unique_media_references(image_sources)
    if not normalized_sources:
        return []

    remove_product_media(product_id)
    stored_paths: list[str] = []
    remote_index = 0
    for source in normalized_sources:
        if is_remote_media_url(source):
            remote_index += 1
            stored_paths.append(
                download_remote_media("products", product_id, source, prefix=f"image-{remote_index}")
            )
            continue
        stored_paths.append(source)
    return stored_paths


def normalize_cell_text(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ")
    return re.sub(r"\s+", " ", text.strip())


def cell_indent(value) -> int:
    if value is None:
        return 0
    text = str(value).replace("\xa0", " ")
    return len(text) - len(text.lstrip(" "))


def load_excel_workbook(file_bytes: bytes):
    return load_workbook(filename=io.BytesIO(file_bytes), read_only=False, data_only=True)


def looks_like_media_reference(value: str) -> bool:
    text = normalize_cell_text(value)
    if not text:
        return False

    parsed = urllib.parse.urlparse(text)
    if parsed.scheme in {"http", "https"} and parsed.netloc:
        return True
    if text.startswith("www."):
        return True
    if text.startswith(("/", "./", "../")):
        return True
    if text.lower().startswith(("media/", "products/", "categories/", "uploads/")):
        return True

    return bool(re.search(r"\.(?:avif|gif|jpe?g|png|svg|webp)(?:\?.*)?$", text, re.IGNORECASE))


def normalize_media_reference(value) -> str:
    text = normalize_cell_text(value)
    if not text:
        return ""
    if text.startswith("www."):
        return f"https://{text}"
    return text


def unique_media_references(values: list[str]) -> list[str]:
    unique: list[str] = []
    seen: set[str] = set()
    for value in values:
        normalized = normalize_media_reference(value)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        unique.append(normalized)
    return unique


def extract_cell_media_candidates(cell) -> list[str]:
    candidates: list[str] = []

    hyperlink = getattr(cell, "hyperlink", None)
    target = getattr(hyperlink, "target", None) or getattr(hyperlink, "location", None)
    if target and looks_like_media_reference(target):
        candidates.append(str(target))

    value = normalize_cell_text(getattr(cell, "value", None))
    if value and looks_like_media_reference(value):
        candidates.append(value)

    return unique_media_references(candidates)


def is_supported_flat_header(header: str) -> bool:
    normalized = normalize_cell_text(header).lower()
    if not normalized:
        return False
    if normalized in FLAT_IMPORT_HEADERS:
        return True
    if normalized in CATEGORY_IMAGE_HEADERS:
        return True
    return bool(PRODUCT_IMAGE_HEADER_RE.match(normalized))


def parse_excel_upload(file_bytes: bytes) -> list[dict[str, str]]:
    workbook = load_excel_workbook(file_bytes)
    sheet = workbook.active

    rows = list(sheet.iter_rows())
    if not rows:
        return []

    header_index = next(
        (
            index
            for index, row in enumerate(rows)
            if any(normalize_cell_text(getattr(cell, "value", None)) for cell in row)
        ),
        None,
    )
    if header_index is None:
        return []

    headers = [normalize_cell_text(getattr(cell, "value", None)).lower() for cell in rows[header_index]]
    if not any(headers):
        return []

    result: list[dict[str, str]] = []
    for row in rows[header_index + 1 :]:
        values = list(row) if row is not None else []
        if not any(normalize_cell_text(getattr(cell, "value", None)) for cell in values):
            continue

        item: dict[str, str] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            cell = values[index] if index < len(values) else None
            if cell is None:
                item[header] = ""
                continue

            if header in CATEGORY_IMAGE_HEADERS or PRODUCT_IMAGE_HEADER_RE.match(header):
                media_candidates = extract_cell_media_candidates(cell)
                item[header] = media_candidates[0] if media_candidates else normalize_media_reference(cell.value)
                continue

            item[header] = normalize_cell_text(cell.value)
        result.append(item)

    return result


def is_flat_excel_import(rows: list[dict[str, str]]) -> bool:
    if not rows:
        return False
    headers = set(rows[0].keys())
    return "title" in headers and any(is_supported_flat_header(header) for header in headers)


def group_consecutive_columns(columns: list[int]) -> list[list[int]]:
    if not columns:
        return []

    groups: list[list[int]] = [[columns[0]]]
    for column in columns[1:]:
        if column == groups[-1][-1] + 1:
            groups[-1].append(column)
        else:
            groups.append([column])
    return groups


def parse_catalog_excel(file_bytes: bytes) -> list[CatalogImportSheet]:
    workbook = load_excel_workbook(file_bytes)
    parsed_sheets: list[CatalogImportSheet] = []

    for sheet in workbook.worksheets:
        rows = [list(row) for row in sheet.iter_rows()]
        used_columns = sorted(
            {
                index + 1
                for row in rows
                for index, cell in enumerate(row)
                if normalize_cell_text(getattr(cell, "value", None)) or extract_cell_media_candidates(cell)
            }
        )
        if not used_columns:
            continue

        root = CatalogImportNode(title=sheet.title, depth=-1, sort=0)
        for block in group_consecutive_columns(used_columns):
            stack: list[CatalogImportNode] = [root]
            block_origin = block[0]

            for row in rows:
                block_cells = []
                for column in block:
                    cell = row[column - 1] if column - 1 < len(row) else None
                    if cell is None:
                        continue
                    cell_text = normalize_cell_text(cell.value)
                    media_candidates = extract_cell_media_candidates(cell)

                    if cell_text and not looks_like_media_reference(cell_text):
                        depth = (column - block_origin) * 100 + cell_indent(cell.value)
                        block_cells.append(
                            {
                                "depth": depth,
                                "title": cell_text,
                                "media_urls": media_candidates,
                            }
                        )
                        continue

                    if media_candidates and block_cells:
                        block_cells[-1]["media_urls"].extend(media_candidates)

                for cell_payload in block_cells:
                    depth = cell_payload["depth"]
                    title = cell_payload["title"]
                    while len(stack) > 1 and stack[-1].depth >= depth:
                        stack.pop()

                    parent = stack[-1]
                    node = CatalogImportNode(
                        title=title,
                        depth=depth,
                        sort=len(parent.children),
                        media_urls=unique_media_references(cell_payload["media_urls"]),
                    )
                    parent.children.append(node)
                    stack.append(node)

        if root.children:
            parsed_sheets.append(CatalogImportSheet(title=sheet.title, nodes=root.children))

    return parsed_sheets


def decimal_or_none(value) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError) as exc:
        raise ValueError("INVALID_DECIMAL") from exc


def bool_from_value(value, *, default: bool = True) -> bool:
    if value in (None, ""):
        return default
    return str(value).strip().lower() not in {"false", "0", "off", "no"}


def ticket_number() -> str:
    return f"TKT-{uuid4().hex[:8].upper()}"


TWOGIS_REVIEW_API_RE = re.compile(
    r"https://public-api\.reviews\.2gis\.com/3\.0/branches/[^\"\\]+"
)


def normalize_2gis_reviews_url(url: str) -> str:
    normalized = (url or "").strip()
    if not normalized:
        raise ValueError("SOURCE_URL_REQUIRED")

    parsed = urllib.parse.urlparse(normalized)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        raise ValueError("INVALID_SOURCE_URL")
    if "2gis.ru" not in parsed.netloc:
        raise ValueError("INVALID_SOURCE_HOST")

    path = parsed.path.rstrip("/")
    if "/firm/" not in path:
        raise ValueError("INVALID_2GIS_FIRM_URL")
    if not path.endswith("/tab/reviews"):
        path = f"{path}/tab/reviews"

    return urllib.parse.urlunparse((parsed.scheme, parsed.netloc, path, "", "", ""))


def fetch_url_text(url: str, *, timeout: int = 20) -> str:
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0 Safari/537.36"
            ),
            "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
        },
    )
    with urllib.request.urlopen(request, timeout=timeout) as response:
        charset = response.headers.get_content_charset() or "utf-8"
        return response.read().decode(charset, errors="ignore")


def extract_2gis_reviews_api_url(page_html: str) -> str:
    match = TWOGIS_REVIEW_API_RE.search(page_html)
    if not match:
        raise ValueError("TWOGIS_REVIEWS_API_NOT_FOUND")
    return match.group(0).replace("offset=50", "offset=0")


def fetch_json(url: str, *, timeout: int = 20) -> dict:
    raw = fetch_url_text(url, timeout=timeout)
    try:
        payload = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise ValueError("INVALID_REMOTE_JSON") from exc
    if not isinstance(payload, dict):
        raise ValueError("INVALID_REMOTE_JSON")
    return payload
